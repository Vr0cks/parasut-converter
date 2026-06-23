"""
Uyumsoft Giden Faturalar → Paraşüt Satış Faturası Dönüştürücü
=============================================================
Kullanım:
    python convert.py
    python convert.py --kaynak "giden faturalar.xlsx" --sablon "parasut_sablon.xlsx" --cikti "yuklenecek.xlsx"
"""

import argparse
import sys
import pandas as pd
from openpyxl import load_workbook

# ── Argümanlar ────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description="Uyumsoft -> Paraşüt fatura dönüştürücü")
parser.add_argument("--kaynak",  default="giden faturalar tümü.xlsx",
                    help="Uyumsoft'tan indirilen Giden Faturalar Excel dosyası")
parser.add_argument("--sablon",  default="parasut_satis_faturalari (2).xlsx",
                    help="Paraşüt'ten indirilen boş şablon Excel dosyası")
parser.add_argument("--cikti",   default="parasut_hazir_yukleme_dosyasi.xlsx",
                    help="Oluşturulacak çıktı dosyasının adı")
args = parser.parse_args()

# ── 1. Uyumsoft verisini oku ──────────────────────────────────────────────────
print(f"[1/4] Uyumsoft dosyasi okunuyor: {args.kaynak}")
try:
    df = pd.read_excel(args.kaynak, engine='openpyxl')
except FileNotFoundError:
    print(f"HATA: '{args.kaynak}' dosyasi bulunamadi.")
    sys.exit(1)

df.columns = df.columns.str.strip()

# Gereken sütunların varlığını kontrol et
zorunlu_sutunlar = ['Fatura No', 'Fatura Tarihi', 'Alıcı', 'Alıcı VKN/TCKN', 'Ödenecek Tutar']
eksik = [s for s in zorunlu_sutunlar if s not in df.columns]
if eksik:
    print(f"HATA: Kaynak dosyada su sutunlar bulunamadi: {eksik}")
    print(f"Mevcut sutunlar: {list(df.columns)}")
    sys.exit(1)

# Fatura No ve VKN/TCKN'deki öncü apostrof/tırnak karakterini temizle
df['Fatura No']        = df['Fatura No'].astype(str).str.lstrip("'").str.strip()
df['Alıcı VKN/TCKN']  = df['Alıcı VKN/TCKN'].astype(str).str.lstrip("'").str.strip()

# Fatura Tarihi: "gg.aa.yyyy SS:dd:ss" → Python datetime nesnesine çevir
def parse_tarih(val):
    try:
        return pd.to_datetime(str(val).strip(), format='%d.%m.%Y %H:%M:%S', dayfirst=True).to_pydatetime()
    except Exception:
        try:
            parsed = pd.to_datetime(str(val).strip(), dayfirst=True, errors='coerce')
            return parsed.to_pydatetime() if pd.notnull(parsed) else None
        except Exception:
            return None

df['_tarih_obj'] = df['Fatura Tarihi'].apply(parse_tarih)
print(f"    {len(df)} fatura okundu.")

# ── 2. Paraşüt şablonunu yükle, örnek satırları temizle ──────────────────────
print(f"[2/4] Parasut sablonu yukleniyor: {args.sablon}")
try:
    wb = load_workbook(args.sablon)
except FileNotFoundError:
    print(f"HATA: '{args.sablon}' sablon dosyasi bulunamadi.")
    print("Parasut > Satis Faturalari > Iceri Aktar ekranindan sablonu indirip bu klasore koyun.")
    sys.exit(1)

ws = wb.active

# Satır 4'ten itibaren tüm mevcut verileri temizle (template örnek satırları)
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        cell.value = None

# ── 3. Verileri Paraşüt şablonuna yaz ────────────────────────────────────────
print("[3/4] Veriler Parasut formatina donusturuluyor...")

# Paraşüt şablonu sütun sırası (satır 3 = başlık satırı):
# 1  MÜŞTERİ ÜNVANI *      | 2  FATURA İSMİ          | 3  FATURA TARİHİ *
# 4  DÖVİZ CİNSİ           | 5  DÖVİZ KURU           | 6  VADE TARİHİ
# 7  TAHSİLAT TL KARŞILIĞI | 8  FATURA TÜRÜ          | 9  FATURA SERİ
# 10 FATURA SIRA NO        | 11 KATEGORİ              | 12 HİZMET/ÜRÜN *
# 13 HİZMET/ÜRÜN AÇIKLAM. | 14 ÇIKIŞ DEPOSU *        | 15 MİKTAR *
# 16 BİRİM FİYATI *        | 17 İNDİRİM TUTARI       | 18 KDV ORANI *
# 19 ÖİV ORANI             | 20 KONAKLAMA VERGİSİ ORANI

hata_sayisi = 0
for i, row in df.iterrows():
    excel_row = i + 4  # 1=yardım metni, 2=boş, 3=başlık → veriler 4. satırdan başlar

    musteri   = str(row.get('Alıcı', '') or '').strip()
    tarih_obj = row['_tarih_obj']
    para_bir  = 'TRL'   # Kaynak dosyada Para Birimi sütunu boş → tümü TRL
    kur       = None    # TRL faturalarda döviz kuru Paraşüt kuralı gereği BOŞ olmalı
    fatura_no = str(row.get('Fatura No', '') or '').strip()
    tutar     = row.get('Ödenecek Tutar', None)
    if pd.isna(tutar) if tutar is not None else False:
        tutar = None
        hata_sayisi += 1

    # Uyumsoft'tan KDV oranları gelmediği için Paraşüt'ün ekstra KDV eklemesini engellemek adına %0 KDV gönderiyoruz.
    # Böylece BİRİM FİYATI direkt Ödenecek Tutar (KDV dahil) oluyor ve toplamlar milimi milimine tutuyor.
    kdvsiz_tutar = tutar

    ws.cell(excel_row,  1, musteri)
    ws.cell(excel_row,  2, fatura_no)   # Uyumsoft fatura kodunu FATURA İSMİ'ne koy
    tarih_cell = ws.cell(excel_row, 3, tarih_obj)
    if tarih_obj:
        tarih_cell.number_format = 'DD-MM-YYYY'
    ws.cell(excel_row,  4, para_bir)
    ws.cell(excel_row,  5, kur)         # TRL → None (Paraşüt zorunlu tutar yabancı dövizde)
    ws.cell(excel_row,  6, None)        # VADE TARİHİ → boş = açık fatura
    ws.cell(excel_row,  7, None)        # TAHSİLAT TL KARŞILIĞI → boş (yalnızca yabancı dövizde)
    ws.cell(excel_row,  8, None)        # FATURA TÜRÜ → boş bırakılırsa "Fatura" kabul edilir
    ws.cell(excel_row,  9, None)        # FATURA SERİ
    ws.cell(excel_row, 10, None)        # FATURA SIRA NO → tam sayı olmalı, Paraşüt otomatik atar
    ws.cell(excel_row, 11, None)        # KATEGORİ
    ws.cell(excel_row, 12, 'Uyumsoft Aktarim Kalemi')
    ws.cell(excel_row, 13, None)        # HİZMET/ÜRÜN AÇIKLAMASI
    ws.cell(excel_row, 14, None)        # ÇIKIŞ DEPOSU
    ws.cell(excel_row, 15, 1)           # MİKTAR
    ws.cell(excel_row, 16, kdvsiz_tutar)# BİRİM FİYATI (KDV dahil gönderiliyor)
    ws.cell(excel_row, 17, None)        # İNDİRİM TUTARI
    ws.cell(excel_row, 18, 0)           # KDV ORANI (%) -> 0 yapıyoruz ki Paraşüt üstüne KDV eklemesin
    ws.cell(excel_row, 19, None)        # ÖİV ORANI
    ws.cell(excel_row, 20, None)        # KONAKLAMA VERGİSİ ORANI

# ── 4. Kaydet ─────────────────────────────────────────────────────────────────
print(f"[4/4] Kaydediliyor: {args.cikti}")
wb.save(args.cikti)

print()
print(f"Tamamlandi!")
print(f"  Toplam fatura : {len(df)}")
if hata_sayisi:
    print(f"  Tutar bos satir: {hata_sayisi} (Parasut'te kontrol edin)")
print(f"  Cikti dosyasi : {args.cikti}")
print(f"  --> Bu dosyayi Parasut > Satis Faturalari > Iceri Aktar ekranindan yukleyin.")