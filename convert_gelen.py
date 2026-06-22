"""
Uyumsoft Gelen Faturalar → Paraşüt Alış Fişi/Faturası Dönüştürücü
==================================================================
Kullanım:
    python convert_gelen.py
    python convert_gelen.py --kaynak "Gelen Faturalar.xlsx" --sablon "parasut_fis_faturalari.xlsx" --cikti "gelen_yuklenecek.xlsx"
"""

import argparse
import sys
import pandas as pd
from openpyxl import load_workbook

# ── Argümanlar ────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description="Uyumsoft Gelen Faturalar -> Paraşüt Alış Faturası dönüştürücü")
parser.add_argument("--kaynak", default="Gelen Faturalar tümü.xlsx",
                    help="Uyumsoft'tan indirilen Gelen Faturalar Excel dosyası")
parser.add_argument("--sablon", default="parasut_fis_faturalari.xlsx",
                    help="Paraşüt'ten indirilen Alış Fişi/Faturası boş şablon dosyası")
parser.add_argument("--cikti",  default="parasut_gelen_hazir.xlsx",
                    help="Oluşturulacak çıktı dosyasının adı")
args = parser.parse_args()

# ── 1. Uyumsoft Gelen Faturalar verisini oku ──────────────────────────────────
print(f"[1/4] Uyumsoft dosyasi okunuyor: {args.kaynak}")
try:
    df = pd.read_excel(args.kaynak, engine='openpyxl')
except FileNotFoundError:
    print(f"HATA: '{args.kaynak}' dosyasi bulunamadi.")
    sys.exit(1)

df.columns = df.columns.str.strip()

# Sütun kontrolü
zorunlu = ['Fatura No', 'Fatura Tarihi', 'Gönderici', 'Ödenecek Tutar']
eksik = [s for s in zorunlu if s not in df.columns]
if eksik:
    print(f"HATA: Kaynak dosyada su sutunlar bulunamadi: {eksik}")
    print(f"Mevcut sutunlar: {list(df.columns)}")
    sys.exit(1)

# Fatura No ve VKN'deki öncü apostrof temizle
df['Fatura No']           = df['Fatura No'].astype(str).str.lstrip("'").str.strip()
df['Gönderici VKN/TCKN'] = df['Gönderici VKN/TCKN'].astype(str).str.lstrip("'").str.strip()

# Fatura Tarihi: "gg.aa.yyyy SS:dd:ss" → Python datetime
def parse_tarih(val):
    try:
        return pd.to_datetime(str(val).strip(), format='%d.%m.%Y %H:%M:%S', dayfirst=True).to_pydatetime()
    except Exception:
        try:
            parsed = pd.to_datetime(str(val).strip(), dayfirst=True, errors='coerce')
            return parsed.to_pydatetime() if pd.notnull(parsed) else None
        except Exception:
            return None

df['_tarih_obj'] = df['Fatura Tarihi'].apply(parse_tarih)
print(f"    {len(df)} fatura okundu.")

# ── 2. Paraşüt şablonunu yükle, örnek satırları temizle ──────────────────────
print(f"[2/4] Parasut sablonu yukleniyor: {args.sablon}")
try:
    wb = load_workbook(args.sablon)
except FileNotFoundError:
    print(f"HATA: '{args.sablon}' sablon dosyasi bulunamadi.")
    print("Parasut > Alis Faturalari > Iceri Aktar ekranindan sablonu indirip bu klasore koyun.")
    sys.exit(1)

ws = wb.active

# Satır 4'ten itibaren tüm mevcut verileri temizle (örnek satırlar)
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        cell.value = None

# ── 3. Verileri Paraşüt şablonuna yaz ────────────────────────────────────────
print("[3/4] Veriler Parasut formatina donusturuluyor...")

# Paraşüt Alış Fişi/Faturası şablonu sütun sırası (satır 3 = başlık):
# 1  Fiş / Fatura Tarihi*
# 2  Tedarikçi
# 3  Kayıt İsmi
# 4  Fiş/Fatura No
# 5  Kategori
# 6  Döviz tipi
# 7  Döviz Kuru      (TRL'de boş)
# 8  Son Ödeme tarihi
# 9  Toplam KDV*
# 10 Toplam Tutar*
# 11 Ödeme Durumu
# 12 Ödeme Hesabı

hata_sayisi = 0
for i, row in df.iterrows():
    excel_row = i + 4  # 1=yardım, 2=boş, 3=başlık → veriler 4'ten başlar

    tarih_obj = row['_tarih_obj']
    tedarikci = str(row.get('Gönderici', '') or '').strip()
    fatura_no = str(row.get('Fatura No', '') or '').strip()
    tutar     = row.get('Ödenecek Tutar', None)
    if pd.isna(tutar) if tutar is not None else False:
        tutar = None
        hata_sayisi += 1

    # Fiş / Fatura Tarihi
    tarih_cell = ws.cell(excel_row, 1, tarih_obj)
    if tarih_obj:
        tarih_cell.number_format = 'DD-MM-YYYY'

    ws.cell(excel_row,  2, tedarikci)   # Tedarikçi (Gönderici)
    ws.cell(excel_row,  3, fatura_no)   # Kayıt İsmi → Uyumsoft fatura kodu
    ws.cell(excel_row,  4, None)        # Fiş/Fatura No → Paraşüt kendi atıyor
    ws.cell(excel_row,  5, 'İçe Aktarım')  # Kategori
    ws.cell(excel_row,  6, 'TRL')       # Döviz tipi
    ws.cell(excel_row,  7, None)        # Döviz Kuru → TRL'de boş
    ws.cell(excel_row,  8, None)        # Son Ödeme tarihi → boş = ödenecek
    ws.cell(excel_row,  9, None)        # Toplam KDV → hesaplanmıyor, Paraşüt halleder
    ws.cell(excel_row, 10, tutar)       # Toplam Tutar (KDV dahil)
    ws.cell(excel_row, 11, 'Ödenecek') # Ödeme Durumu
    ws.cell(excel_row, 12, None)        # Ödeme Hesabı → boş

# ── 4. Kaydet ─────────────────────────────────────────────────────────────────
print(f"[4/4] Kaydediliyor: {args.cikti}")
wb.save(args.cikti)

print()
print(f"Tamamlandi!")
print(f"  Toplam fatura : {len(df)}")
if hata_sayisi:
    print(f"  Tutar bos satir: {hata_sayisi} (Parasut'te kontrol edin)")
print(f"  Cikti dosyasi : {args.cikti}")
print(f"  --> Bu dosyayi Parasut > Alis Faturalari > Iceri Aktar ekranindan yukleyin.")
