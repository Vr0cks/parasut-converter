import os
import sys
import time
import requests
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
from datetime import datetime

# Load Environment Variables
load_dotenv()
CLIENT_ID = os.getenv('PARASUT_CLIENT_ID')
CLIENT_SECRET = os.getenv('PARASUT_CLIENT_SECRET')
USERNAME = os.getenv('PARASUT_USERNAME')
PASSWORD = os.getenv('PARASUT_PASSWORD')
COMPANY_ID = os.getenv('PARASUT_COMPANY_ID')

class ParasutAPI:
    def __init__(self):
        self.token = self._get_token()
        self.headers = {"Authorization": f"Bearer {self.token}"}
        
    def _get_token(self):
        if not all([CLIENT_ID, CLIENT_SECRET, USERNAME, PASSWORD, COMPANY_ID]):
            print("\n[HATA] .env dosyası eksik veya API bilgileri girilmemiş!")
            print("Lütfen .env dosyasını oluşturun ve bilgilerinizi girin.")
            sys.exit(1)
            
        print("Paraşüt API'ye bağlanılıyor...")
        res = requests.post(
            'https://api.parasut.com/oauth/token', 
            data={'username': USERNAME, 'password': PASSWORD, 'grant_type': 'password', 'redirect_uri': 'urn:ietf:wg:oauth:2.0:oob'}, 
            auth=(CLIENT_ID, CLIENT_SECRET)
        )
        if res.status_code != 200:
            print("[HATA] API'ye bağlanılamadı. Bilgilerinizi kontrol edin.")
            sys.exit(1)
        return res.json()['access_token']

    def fetch_all(self, endpoint):
        results = []
        url = f"https://api.parasut.com/v4/{COMPANY_ID}/{endpoint}?page[size]=25"
        while url:
            res = requests.get(url, headers=self.headers)
            if res.status_code == 429:
                time.sleep(2)
                continue
            data = res.json()
            results.extend(data.get('data', []))
            url = data.get('links', {}).get('next')
        return results

    def delete_record(self, endpoint, record_id):
        url = f"https://api.parasut.com/v4/{COMPANY_ID}/{endpoint}/{record_id}"
        while True:
            res = requests.delete(url, headers=self.headers)
            if res.status_code == 429:
                time.sleep(1)
                continue
            return res.status_code == 204

    def update_record(self, endpoint, record_id, payload):
        url = f"https://api.parasut.com/v4/{COMPANY_ID}/{endpoint}/{record_id}"
        while True:
            res = requests.put(url, headers=self.headers, json=payload)
            if res.status_code == 429:
                time.sleep(1)
                continue
            return res.status_code == 200


def clean_balance(val):
    if pd.isna(val): return 0.0
    if isinstance(val, (int, float)): return float(val)
    val_str = str(val).replace('.', '').replace(',', '.')
    try: return float(val_str)
    except: return 0.0

def normalize_name(name):
    return str(name).strip().lower().replace(' ', '').replace('i̇', 'i').replace('ı', 'i')


def convert_invoices():
    print("\n--- Uyumsoft Faturalarını Paraşüt Şablonuna Çevirme ---")
    print("1. Satış Faturalarını Çevir (convert.py)")
    print("2. Alış Faturalarını Çevir (convert_gelen.py)")
    secim = input("Seçiminiz: ")
    
    if secim == '1':
        if os.path.exists("convert.py"):
            print("Çalıştırılıyor: convert.py...")
            os.system("python convert.py")
        else:
            print("[HATA] convert.py bulunamadı.")
    elif secim == '2':
        if os.path.exists("convert_gelen.py"):
            print("Çalıştırılıyor: convert_gelen.py...")
            os.system("python convert_gelen.py")
        else:
            print("[HATA] convert_gelen.py bulunamadı.")
    else:
        print("İptal edildi.")

def delete_all_invoices(api):
    print("\n--- Tüm Faturaları Toplu Sil (Dikkat: Geri Alınamaz!) ---")
    onay = input("Tüm SATIŞ ve ALIŞ faturalarını silmek istediğinize emin misiniz? (E/H): ")
    if onay.lower() != 'e':
        print("İptal edildi.")
        return
        
    for endpoint in ['sales_invoices', 'purchase_bills']:
        print(f"\n{endpoint} çekiliyor...")
        invoices = api.fetch_all(endpoint)
        print(f"Toplam {len(invoices)} adet fatura bulundu.")
        
        deleted = 0
        for inv in invoices:
            if api.delete_record(endpoint, inv['id']):
                deleted += 1
                print(f"Silindi: {inv['id']} ({deleted}/{len(invoices)})", end='\r')
        print(f"\n{endpoint} için silme işlemi tamamlandı.")

def delete_duplicates(api):
    print("\n--- Mükerrer (Çift) Faturaları Temizleme ---")
    invoices = api.fetch_all('sales_invoices')
    print(f"Toplam {len(invoices)} adet satış faturası çekildi.")
    
    seen = {}
    duplicates = []
    
    for inv in invoices:
        attrs = inv['attributes']
        key = (attrs.get('net_total'), attrs.get('issue_date'))
        if key in seen:
            duplicates.append(inv)
        else:
            seen[key] = inv
            
    print(f"{len(duplicates)} adet çift fatura tespit edildi.")
    if not duplicates: return
    
    onay = input("Mükerrer faturaları silmek istiyor musunuz? (E/H): ")
    if onay.lower() == 'e':
        for i, dup in enumerate(duplicates):
            api.delete_record('sales_invoices', dup['id'])
            print(f"Siliniyor... {i+1}/{len(duplicates)}", end='\r')
        print("\nÇift faturalar temizlendi!")

def update_vkns(api):
    print("\n--- Uyumsoft'tan Eksik VKN ve Adresleri Çekip Güncelleme ---")
    try:
        df = pd.read_excel('cari uyumsoft.xlsx', engine='openpyxl')
    except FileNotFoundError:
        print("[HATA] 'cari uyumsoft.xlsx' bulunamadı!")
        return
        
    contacts = api.fetch_all('contacts')
    print(f"Paraşüt'ten {len(contacts)} cari çekildi.")
    
    df.columns = df.columns.str.strip()
    uyumsoft_dict = {}
    for _, row in df.iterrows():
        name = normalize_name(row.get('İsim / Ünvan', ''))
        vkn = str(row.get('Vergi Numarası / Tc Kimlik No', '')).strip()
        vergi_dairesi = str(row.get('Vergi Dairesi', '')).strip()
        
        # 'nan' değerleri temizle
        if vkn.lower() == 'nan': vkn = ''
        if vergi_dairesi.lower() == 'nan': vergi_dairesi = ''
        
        if name:
            uyumsoft_dict[name] = {'vkn': vkn, 'vergi_dairesi': vergi_dairesi}
            
    updated = 0
    for c in contacts:
        attrs = c['attributes']
        c_name = normalize_name(attrs.get('name', ''))
        
        if c_name in uyumsoft_dict:
            u_data = uyumsoft_dict[c_name]
            payload = {"data": {"id": c['id'], "type": "contacts", "attributes": {}}}
            needs_update = False
            
            if u_data['vkn'] and not attrs.get('tax_number') and not attrs.get('tc_id'):
                if len(u_data['vkn']) == 11: payload['data']['attributes']['tc_id'] = u_data['vkn']
                else: payload['data']['attributes']['tax_number'] = u_data['vkn']
                needs_update = True
                
            if u_data['vergi_dairesi'] and not attrs.get('tax_office'):
                payload['data']['attributes']['tax_office'] = u_data['vergi_dairesi']
                needs_update = True
                
            if needs_update:
                if api.update_record('contacts', c['id'], payload):
                    updated += 1
                    
    print(f"İşlem Tamam! {updated} adet müşterinin bilgileri güncellendi.")

def sync_balances(api):
    print("\n--- Bakiyeleri Karşılaştır ve Düzeltme Faturası Üret ---")
    try:
        df = pd.read_excel('cari uyumsoft.xlsx', engine='openpyxl')
    except FileNotFoundError:
        print("[HATA] 'cari uyumsoft.xlsx' bulunamadı!")
        return

    contacts = api.fetch_all('contacts')
    parasut_map = {}
    for c in contacts:
        attrs = c['attributes']
        norm_name = normalize_name(attrs.get('name', ''))
        parasut_map[norm_name] = {'name': attrs.get('name', ''), 'balance': float(attrs.get('trl_balance', 0) or 0)}

    df.columns = df.columns.str.strip()
    satis_listesi, alis_listesi = [], []
    
    for _, row in df.iterrows():
        u_name = row.get('İsim / Ünvan', '')
        if not u_name or str(u_name) == 'nan': continue
        
        u_bal = clean_balance(row.get('Açık Bakiye', 0))
        norm_uname = normalize_name(u_name)
        
        p_bal = parasut_map.get(norm_uname, {}).get('balance', 0.0)
        p_name = parasut_map.get(norm_uname, {}).get('name', u_name)
        
        diff = u_bal - p_bal
        if abs(diff) > 1.0:
            if diff > 0: satis_listesi.append({'name': p_name, 'amount': diff})
            else: alis_listesi.append({'name': p_name, 'amount': abs(diff)})
            
    print(f"Satış Faturası kesilecek cari sayısı: {len(satis_listesi)}")
    print(f"Alış Faturası kesilecek cari sayısı: {len(alis_listesi)}")
    
    # Excel dosyalarını oluştur (Mevcut mantık ile)
    tarih_obj = datetime.now()
    if satis_listesi and os.path.exists("parasut_satis_faturalari (2).xlsx"):
        wb = load_workbook("parasut_satis_faturalari (2).xlsx")
        ws = wb.active
        for r in range(4, 10):
            for c in range(1, 21): ws.cell(r, c, None)
        for i, item in enumerate(satis_listesi):
            row = 4 + i
            ws.cell(row, 1, item['name'])
            ws.cell(row, 2, "Geçmiş Dönem Bakiye Düzeltmesi")
            ws.cell(row, 3, tarih_obj).number_format = 'DD-MM-YYYY'
            ws.cell(row, 4, "TRL")
            ws.cell(row, 8, "Fatura")
            ws.cell(row, 9, "DZT")
            ws.cell(row, 10, i + 1)
            ws.cell(row, 11, "Devir")
            ws.cell(row, 12, "Bakiye Düzeltme")
            ws.cell(row, 15, 1)
            ws.cell(row, 16, item['amount'])
            ws.cell(row, 17, 0)
            ws.cell(row, 18, 0)
        wb.save("bakiye_duzeltme_satis.xlsx")
        print("bakiye_duzeltme_satis.xlsx oluşturuldu.")
        
    if alis_listesi and os.path.exists("parasut_fis_faturalari.xlsx"):
        wb = load_workbook("parasut_fis_faturalari.xlsx")
        ws = wb.active
        for r in range(4, 10):
            for c in range(1, 13): ws.cell(r, c, None)
        for i, item in enumerate(alis_listesi):
            row = 4 + i
            ws.cell(row, 1, tarih_obj).number_format = 'DD-MM-YYYY'
            ws.cell(row, 2, item['name'])
            ws.cell(row, 3, "Geçmiş Dönem Bakiye Düzeltmesi")
            ws.cell(row, 4, f"DZT-{i+1}")
            ws.cell(row, 5, "Devir")
            ws.cell(row, 6, "TRL")
            ws.cell(row, 9, 0)
            ws.cell(row, 10, item['amount'])
            ws.cell(row, 11, "Ödenecek")
        wb.save("bakiye_duzeltme_alis.xlsx")
        print("bakiye_duzeltme_alis.xlsx oluşturuldu.")


def main_menu():
    print("="*50)
    print("    PARAŞÜT GÖÇ VE YÖNETİM ASİSTANI V1.0")
    print("="*50)
    
    api = ParasutAPI()
    
    while True:
        print("\n[ MENÜ ]")
        print("1. Uyumsoft Faturalarını Dönüştür (Yakında)")
        print("2. Tüm Faturaları Toplu Sil (Sıfırlama)")
        print("3. Mükerrer (Çift) Faturaları Temizle")
        print("4. Uyumsoft'tan Eksik VKN ve Adresleri Çekip Güncelle")
        print("5. Bakiyeleri Karşılaştır ve Düzeltme Faturası Üret")
        print("0. Çıkış")
        
        secim = input("Lütfen bir işlem seçin (0-5): ")
        
        if secim == '0':
            print("Çıkış yapılıyor. İyi çalışmalar!")
            break
        elif secim == '1': convert_invoices()
        elif secim == '2': delete_all_invoices(api)
        elif secim == '3': delete_duplicates(api)
        elif secim == '4': update_vkns(api)
        elif secim == '5': sync_balances(api)
        else: print("Geçersiz seçim, tekrar deneyin.")

if __name__ == "__main__":
    main_menu()
